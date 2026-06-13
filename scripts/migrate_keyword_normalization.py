# -*- coding: utf-8 -*-
"""
final의 '정규화' 컬럼을 신규 생성 파일에 이식하고
keyword_frequency_review_final.xlsx 로 저장
"""
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE = Path("data/processed")
SRC_NEW   = BASE / "keyword_channel_frequency_review.xlsx"
SRC_FINAL = BASE / "keyword_channel_frequency_review_final.xlsx"
OUT       = BASE / "keyword_frequency_review_final.xlsx"


def main():
    new   = pd.read_excel(SRC_NEW)
    final = pd.read_excel(SRC_FINAL)

    # final 정규화 매핑 구성
    norm_map = (
        final.dropna(subset=["정규화"])
             .set_index("키워드")["정규화"]
             .to_dict()
    )
    print(f"final 정규화 매핑: {len(norm_map)}개")

    # 신규 파일에 이식
    before = new["정규화"].notna().sum()
    new["정규화"] = new["키워드"].map(norm_map)
    after  = new["정규화"].notna().sum()
    print(f"이식 완료: {before} → {after}개 ({after - before:+d})")
    print(f"빈칸 유지 (신규 키워드): {new['정규화'].isna().sum()}개")

    # 스타일링 저장
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    norm_fill   = PatternFill("solid", fgColor="E2EFDA")  # 연초록 - 정규화 채워진 행

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        new.to_excel(writer, index=False, sheet_name="keyword_frequency")
        ws = writer.book["keyword_frequency"]

        # 컬럼 너비
        col_widths = {
            "키워드":           24,
            "블로그 등장빈도":  14,
            "블로그_제품":      30,
            "인스타그램 등장빈도": 16,
            "인스타_제품":      30,
            "IP 등장빈도":      12,
            "IP_목록":          30,
            "트렌드 등장빈도":  14,
            "트렌드_소스키워드": 30,
            "정규화":           30,
        }
        for col_idx, col_name in enumerate(new.columns, start=1):
            width = col_widths.get(col_name, 16)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 헤더 스타일
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # 정규화 채워진 행 하이라이트
        norm_col_idx = list(new.columns).index("정규화") + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            norm_cell = ws.cell(row=row_idx, column=norm_col_idx)
            if norm_cell.value is not None and str(norm_cell.value).strip():
                for cell in row:
                    cell.fill = norm_fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"\n저장 완료: {OUT}")
    print(f"총 {len(new)}행 / 정규화 {after}개 채움")


if __name__ == "__main__":
    main()
